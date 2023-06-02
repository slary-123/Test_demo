
# 001是一个户型随机匹配某个方案
from openpyxl.drawing.image import Image
from common.Base_options import Base_options as bo
from testdata.locators import locators as loc
import pyautogui as pg
import os,time,pyperclip,openpyxl,random
from PIL import Image
from common.Doexcel import Doexcel
from common.my_logger import logger
from itertools import cycle
class  Test_index_page:
    def test_upload(self):
        count_url = 0
        bo().move(loc.index_page.huxing)
        bo().click_loc(loc.index_page.daorutupian)
        faillist = []
        count = 1
        wb = openpyxl.load_workbook(r'C:\Users\User\Desktop\2222\1.xlsx')
        print("调试1")
        ws = wb['Sheet1']
        ws = wb.active
        ws.cell(1, 1, "方案名称")  # 指定行列进行追加
        ws.cell(1, 2, "大师方案名称")
        ws.cell(1, 3, "原始户型图")
        ws.cell(1, 4, "户型识别图")
        ws.cell(1, 5, "地面截图")
        ws.cell(1, 6, "顶面截图")
        ws.cell(1, 7, "整屋全景图")
        print("调试2")
        # 使用 `os.walk()` 方法遍历目录树，返回三个值：P:当前目录路径、m:当前目录下的子目录列表、当前目录下的文件列表。
        for p, m, names in os.walk(loc.index_page.path): # 遍历文件
            print("当前目录路径:",p)
            print("当前目录下的文件列表:", names)
            print('---户型图总数：', len(names))
            for name in names:
                print('---户型id：', name)
                try:
                    file = os.path.join(loc.index_page.path,name)
                    print("上传图片路径:",file)
                    bo().upload(file)  # 调用上传图片
                    time.sleep(2)
                    bo().click_loc(loc.index_page.zidongshibie)
                    time.sleep(7)
                    for i in range(20):
                        if not bo().get_loc(loc.index_page.xin_jian_fang_an):
                            print("正在生成户型test...")
                            time.sleep(5)
                        elif bo().get_loc(loc.index_page.xin_jian_fang_an):
                            image = Image.open(file)
                            shotfile_sb = os.path.join(loc.index_page.screenshotpath, f"sb_{name}")
                            print("shotfile_sb的图片路径：", shotfile_sb)
                            pg.screenshot(shotfile_sb)  # 截图保存到shotfile文件路径中
                            print(f"-----{name} 截图完成")
                            time.sleep(2)
                            bo().click_loc(loc.index_page.x)
                            bo().click_loc(loc.index_page.baocun_tuichu)
                            break

                    time.sleep(20)
                    pg.click(x=391, y=253) #点击原始户型图
                    print("点击原始户型图")
                    for i in range(20):
                        if not bo().get_loc(loc.index_page.xingmu_list):
                            pg.click(x=391, y=253)  # 点击原始户型图
                        if bo().get_loc(loc.index_page.xingmu_list):
                            bo().click_loc(loc.index_page.weimingming)
                            pyperclip.copy(name)
                            pg.hotkey('ctrl', 'v')
                            pg.press('enter')
                            break


                    #list = ['浪漫满屋-北欧','中古之风，沉淀之美','不被定义的生活-中古','沐·白-奶油','奢华欧式-璀目眩烂','墨轩—现代风格','都市菁英的 “礼与韵”-后现代','轻奢-时光之境','现代-轻盈之家','诧寂-夜幕之美','白木现代-干净慵懒','法式-治愈温馨','北欧-暖心居所']
                    #for list_i in list:
                    while True:
                        list = ['北欧-质雅温润', 'L形衣柜卧室', '法式带儿童房案例', '现代黑白灰，静谧之家',
                                '日式-原木风', '轻奢-宁静之美', '优雅欧式-浪漫之家', '古朴禅意、诧寂风',
                                '中式-简洁端庄', '自然清新、北欧风']
                        list_i = random.choice(list)
                        bo().click_loc(loc.index_page.xinjianfangan)
                        time.sleep(7)
                        bo().click_loc(loc.index_page.anli_name)
                        logger.info(list_i)  #"方案名称：",f"{name},_,{list_i}"
                        pyperclip.copy(list_i) # 复制到剪贴板
                        pg.hotkey('ctrl', 'v')
                        pg.press('enter')
                        time.sleep(2)
                        pg.moveTo(x=241, y=380)  # 移动至大师方案图片上
                        pg.click()  # 点击大师方案图片上的应用按钮
                        time.sleep(3)
                        pg.moveTo(x=1421, y=834)  # 移动至一键应用上
                        pg.click()  # 点击一键应用
                        for i in range(50):
                            if not bo().get_loc(loc.index_page.shihui_moxing):
                                pg.moveTo(x=241, y=380)  # 移动其他地方
                                pg.moveTo(x=1421, y=834)  # 移动至一键应用
                                pg.click()
                                time.sleep(3)
                            if bo().get_loc(loc.index_page.shihui_moxing):
                                pg.moveTo(x=1470, y=555)  # 点击知道了
                                pg.click()
                                break
                        for i in range(10):
                            if not bo().get_loc(loc.index_page.yinngyzhidaole):  # 知道了弹窗的 应用
                                break
                            if bo().get_loc(loc.index_page.yinngyzhidaole):  # 知道了弹窗的 应用
                                bo().click_loc(loc.index_page.zhidaole)
                                pg.moveTo(x=1470, y=555)
                                pg.click()  # 点击知道了
                            else:
                                break
                        time.sleep(3)
                        pg.typewrite(['1'])
                        time.sleep(2)
                        dimian="地面"
                        print("各名称：",f"name={name},list_i={list_i},dimian={dimian}.png")
                        shotfile_dimian = os.path.join(loc.index_page.screenshotpath, f"{name},{list_i},{dimian}.png")
                        print("地面图片路径：",shotfile_dimian)
                        pg.screenshot(shotfile_dimian)  # 截图保存到shotfile文件路径中
                        print(f"-----{name} 截图完成_地面")
                        time.sleep(3)
                        pg.typewrite(['2'])
                        time.sleep(2)
                        dingmian = "顶面"
                        shotfile_dingmian = os.path.join(loc.index_page.screenshotpath, f"{name},{list_i},{dingmian}.png")
                        print("顶面图片路径：",shotfile_dingmian)
                        pg.screenshot(shotfile_dingmian)  # 截图保存到shotfile文件路径中
                        print(f"-----{name} 截图完成_顶面")
                        time.sleep(3)
                        pg.typewrite(['3'])
                        time.sleep(3)
                        bo().move(loc.index_page.daochu)
                        bo().click_loc(loc.index_page.zheng_quan)
                        bo().click_loc(loc.index_page.shengcheng)
                        pg.moveTo(x=1022, y=906)
                        pg.moveTo(x=1022, y=706)  # 生成按钮
                        pg.click(x=1022, y=706) # 生成按钮
                        for i in range(10):
                            if not bo().get_loc(loc.index_page.quan_jing_tu):  # 全景图生成进度
                                bo().click_loc(loc.index_page.shengcheng)  # 生成按钮
                                pg.moveTo(x=1022, y=906)
                                pg.moveTo(x=1022, y=706)# 生成按钮
                                pg.click(x=1022, y=706)
                            else:
                                break
                        for i in range(70):
                            if not bo().get_loc(loc.index_page.quanjingtu_ok):
                                bo().click_loc(loc.index_page.shengcheng)
                                pg.moveTo(x=1022, y=706) # 生成按钮
                                pg.click(x=1022, y=706)
                                time.sleep(1)
                            if bo().get_loc(loc.index_page.quanjingtu_ok):
                                bo().click_loc(loc.index_page.fuzhi)
                                text = pyperclip.paste() #从剪贴板中获取文本内容（text）
                                count_url += 1
                                print("复制链接地址", text)
                          #      Doexcel(r"C:\Users\User\Pictures\3333\3.xlsx", "Sheet1").weite_back(count_url + 1,name,list_i,text)
                                time.sleep(2)

                                count += 1
                                # 写入数据

                                ws.cell(count, 1, f"{name}_{list_i}")  # 指定行列进行追加
                                ws.cell(count, 2, list_i)
                                ws.cell(count, 7, text)

                                ws.row_dimensions[count].height = 80  # ws对象.row_dimensions[行]高=值
                                ws.column_dimensions['A'].width = 13 #设置第A列的列宽
                                ws.column_dimensions['B'].width = 16
                                ws.column_dimensions['C'].width = 18
                                ws.column_dimensions['D'].width = 22
                                ws.column_dimensions['E'].width = 22
                                ws.column_dimensions['F'].width = 22
                                ws.column_dimensions['G'].width = 40
                                img_C = openpyxl.drawing.image.Image(file)
                                print("img=：", img_C)
                                img_C.width = 114  # 设置图片的宽度
                                img_C.height = 101  # 设置图片的高度
                                ws.add_image(img_C, 'C{}'.format(count))
                                img_C = openpyxl.drawing.image.Image(file)

                                img_D = openpyxl.drawing.image.Image(shotfile_sb)
                                img_D.width = 150  # 设置图片的宽度
                                img_D.height = 101  # 设置图片的高度
                                ws.add_image(img_D, 'D{}'.format(count))
                                img_D = openpyxl.drawing.image.Image(shotfile_sb)

                                print("调试shotfile_dimian的路径：",shotfile_dimian)
                                img_E = openpyxl.drawing.image.Image(shotfile_dimian)
                                img_E.width = 150  # 设置图片的宽度
                                img_E.height = 101  # 设置图片的高度
                                ws.add_image(img_E, 'E{}'.format(count))
                                img_E = openpyxl.drawing.image.Image(shotfile_dimian)

                                img_F = openpyxl.drawing.image.Image(shotfile_dingmian)
                                print("img=：", img_F)
                                img_F.width = 150  # 设置图片的宽度
                                img_F.height = 101  # 设置图片的高度
                                ws.add_image(img_F, 'F{}'.format(count))
                                img_F = openpyxl.drawing.image.Image(shotfile_dingmian)

                                wb.save(r"C:\Users\User\Pictures\3333\20230711\10.xlsx")
                                ws = wb.active

                                print('程序结束！')
                                break
                        time.sleep(3)
                        pg.moveTo(x=1142, y=742)
                        pg.click()
                        bo().click_loc(loc.index_page.quxiao_01)
                        for i in range(50):
                            if bo().get_loc(loc.index_page.sheng_cheng_zheng):
                                bo().click_loc(loc.index_page.quxiao_01)
                                pg.moveTo(x=1142, y=742)
                                pg.click()
                                pg.moveTo(x=1342, y=842)
                            else:
                                break
                        bo().click_loc(loc.index_page.x)
                        bo().click_loc(loc.index_page.baocun_tuichu)
                        time.sleep(20)
                        if bo().get_loc(loc.index_page.xiangmu_guanli):
                            bo().click_loc(loc.index_page.weimingmingfangan)
                            time.sleep(2)
                            pyperclip.copy(f"{name}_{list_i}")
                            time.sleep(2)
                            pg.hotkey('ctrl', 'v')
                            time.sleep(2)
                            pg.press('enter')
                        if not bo().get_loc(loc.index_page.weimingmingfangan):
                            break
                        break
                    #     for x in range(10):
                    #         if bo().get_loc(loc.create_plan.xiangmu_guanli):
                    #             for i in range(10):
                    #                 time.sleep(2)
                    #                 if bo().get_loc(loc.create_plan.weimingmingfangan):
                    #                     bo().click_loc(loc.create_plan.weimingmingfangan)
                    #                     time.sleep(2)
                    #                     pyperclip.copy(f"{name}_{list_i}")
                    #                     time.sleep(2)
                    #                     pg.hotkey('ctrl', 'v')
                    #                     time.sleep(2)
                    #                     pg.press('enter')
                    #                 if not bo().get_loc(loc.create_plan.weimingmingfangan):
                    #                     break
                    #             break
                    # time.sleep(5)
                    bo().click_loc(loc.index_page.xiangmu_guanli)
                    bo().click_loc(loc.index_page.xinjian)
                    for i in range(10):
                        if not bo().get_loc(loc.index_page.ziyouhuizhi):
                            bo().click_loc(loc.index_page.xinjian)
                        else:
                            break
                    for i in range(10):
                        if not bo().get_loc(loc.index_page.huxing):
                            bo().click_loc(loc.index_page.ziyouhuizhi)
                        else:
                            break
                    bo().move(loc.index_page.huxing)
                    bo().click_loc(loc.index_page.daorutupian)
                except Exception:
                    logger.info(f"{name[:-4]} 执行失败了！")
                    faillist.append(name[:-5])
        logger.info('执行完毕')