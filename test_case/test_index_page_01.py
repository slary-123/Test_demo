from openpyxl.drawing.image import Image
from common.Base_options import Base_options as bo
from testdata.locators import locator as loc
import pyautogui as pg
import os,time,pyperclip,openpyxl
from PIL import Image
from common.Doexcel import Doexcel
from common.my_logger import logger
from itertools import cycle
print()
"""
一个户型，循环每个方案
"""
class  Test_index_page:
    def test_upload(self): # 测试上传
        count_url = 0  # url数为0
        bo().move(loc.index_page.huxing) # 移动到户型，hover显示
        bo().click_loc(loc.index_page.daorutupian) # 点击导入图片
        faillist = [] # 定义一个失败列表，初始为空
        count = 1 # 定义一个计数器
        wb = openpyxl.load_workbook(r'C:\Users\user\Desktop\sheet-file\plan-house.xlsx') # 通过文件得到一个工作薄，参数是文件名，如果有路径要写绝对路径
        print("调试1") # 打断点
        ws = wb['Sheet1'] # 获取sheet1表对象
        ws = wb.active # 当前活动工作表对应的worksheet对象
        ws.cell(1, 1, "方案名称")  # 指定行列进行追加调用的源码是：def cell(self, row, column, value=None):
        ws.cell(1, 2, "大师方案名称")
        ws.cell(1, 3, "原始户型图")
        ws.cell(1, 4, "户型识别图")
        ws.cell(1, 5, "地面截图")
        ws.cell(1, 6, "顶面截图")
        ws.cell(1, 7, "整屋全景图")
        print("调试2") # 打断点
        # 使用 `os.walk()` 方法遍历目录树，返回三个值：root:当前根目录路径、dirs:当前根目录下的子目录列表、files:当前子目录下的文件列表。
        for root, dirs, files in os.walk(loc.index_page.path): # loc.create_plan.path= r"C:\Users\User\Desktop\big-house" 放户型图的文件
            print("根目录路径:",root)
            print("子目录路径:",dirs)
            print("当前目录下的文件列表:", files)
            print('---户型图总数：', len(files))
            for name in files: # 遍历列表
                print('---户型id：', name[:-4])
                try:
                    file = os.path.join(loc.index_page.path,name)
                    print("上传图片路径:",file)
                    bo().upload(file)  # 调用上传图片
                    bo().click_loc(loc.index_page.zidongshibie)
                    time.sleep(7)
                    for i in range(20):
                        if not bo().get_loc(loc.index_page.xin_jian_fang_an):
                            print("正在生成户型test...")
                            time.sleep(5)
                        elif bo().get_loc(loc.index_page.xin_jian_fang_an):
                            image = Image.open(file)
                            shotfile_sb = os.path.join(loc.index_page.screenshotpath, f"sb_{name}") # 户型识别之后的截图
                            print("shotfile_sb的图片路径：", shotfile_sb)
                            pg.screenshot(shotfile_sb)  # 截图保存到shotfile文件路径中
                            print(f"-----{name} 截图完成")
                            time.sleep(2)
                            bo().click_loc(loc.index_page.x)
                            bo().click_loc(loc.index_page.baocun_tuichu)
                            break
                    time.sleep(3)
                    pg.click(x=391, y=253) #点击原始户型图
                    for i in range(20):
                        if not bo().get_loc(loc.index_page.xingmu_list):
                            pg.click(x=391, y=253)  # 点击原始户型图
                        if bo().get_loc(loc.index_page.xingmu_list):
                            bo().click_loc(loc.index_page.weimingming)
                            pyperclip.copy(name)
                            pg.hotkey('ctrl', 'v')
                            pg.press('enter')
                            break
                    count_test = 1
                    list = ['浪漫满屋-北欧','中古之风，沉淀之美']
                    for list_i in list:
                        bo().click_loc(loc.index_page.xinjianfangan) # 新建方案
                        time.sleep(7)
                        bo().click_loc(loc.index_page.anli_name)
                        logger.info(list_i)  #"方案名称：",f"{name},_,{list_i}"
                        pyperclip.copy(list_i) # 复制到剪贴板
                        pg.hotkey('ctrl', 'v')
                        pg.press('enter')
                        time.sleep(2)
                        pg.moveTo(x=241, y=380)  # 移动至大师方案图片上
                        pg.click()  # 点击大师方案图片上的应用按钮
                        time.sleep(3)
                        pg.moveTo(x=1421, y=834)  # 移动至一键应用上
                        pg.click()  # 点击一键应用
                        for i in range(50):
                            if not bo().get_loc(loc.index_page.shihui_moxing):
                                pg.moveTo(x=241, y=380)  # 移动其他地方
                                pg.moveTo(x=1421, y=834)  # 移动至一键应用
                                pg.click()
                                time.sleep(3)
                            if bo().get_loc(loc.index_page.shihui_moxing):
                                pg.moveTo(x=1470, y=555)  # 点击知道了
                                pg.click()
                                break
                        for i in range(10):
                            if not bo().get_loc(loc.index_page.yinngyzhidaole):  # 知道了弹窗的 应用
                                break
                            if bo().get_loc(loc.index_page.yinngyzhidaole):  # 知道了弹窗的 应用
                                bo().click_loc(loc.index_page.zhidaole)
                                pg.moveTo(x=1470, y=555)
                                pg.click()  # 点击知道了
                            else:
                                break
                        time.sleep(3)
                        pg.typewrite(['1'])
                        time.sleep(2)
                        dimian="地面"
                        shotfile_dimian = os.path.join(loc.index_page.screenshotpath, f"{name},{count_test},{list_i},{dimian}.png")
                        pg.screenshot(shotfile_dimian)  # 截图保存到shotfile文件路径中
                        logger.info(f"-----{name} 截图完成")
                        time.sleep(3)
                        pg.typewrite(['2'])
                        time.sleep(2)
                        dingmian = "顶面"
                        shotfile_dingmian = os.path.join(loc.index_page.screenshotpath, f"{name},{count_test},{list_i},{dingmian}.png")
                        pg.screenshot(shotfile_dingmian)  # 截图保存到shotfile文件路径中
                        logger.info(f"-----{name} 截图完成")
                        time.sleep(3)
                        pg.typewrite(['3'])
                        time.sleep(3)
                        bo().move(loc.index_page.daochu)
                        bo().click_loc(loc.index_page.zheng_quan)
                        bo().click_loc(loc.index_page.shengcheng)
                        pg.moveTo(x=1022, y=906)
                        pg.moveTo(x=1022, y=706)  # 生成按钮
                        pg.click(x=1022, y=706) # 生成按钮
                        for i in range(10):
                            if not bo().get_loc(loc.index_page.quan_jing_tu):  # 全景图生成进度
                                bo().click_loc(loc.index_page.shengcheng)  # 生成按钮
                                pg.moveTo(x=1022, y=906)
                                pg.moveTo(x=1022, y=706)# 生成按钮
                                pg.click(x=1022, y=706)
                            else:
                                break
                        for i in range(70):
                            if not bo().get_loc(loc.index_page.quanjingtu_ok):
                                bo().click_loc(loc.index_page.shengcheng)
                                pg.moveTo(x=1022, y=706) # 生成按钮
                                pg.click(x=1022, y=706)
                                time.sleep(1)
                            if bo().get_loc(loc.index_page.quanjingtu_ok):
                                bo().click_loc(loc.index_page.fuzhi)
                                text = pyperclip.paste() #从剪贴板中获取文本内容（text）
                                count_url += 1
                                print("复制链接地址", text)
                          #      Doexcel(r"C:\Users\User\Pictures\3333\3.xlsx", "Sheet1").weite_back(count_url + 1,name,list_i,text)
                                time.sleep(2)
                                count += 1
                                # 写入数据
                                ws.cell(count, 1, f"{name}_{list_i}")  # 指定行列进行追加
                                ws.cell(count, 2, list_i)
                                ws.cell(count, 7, text)
                                ws.row_dimensions[count].height = 80  # ws对象.row_dimensions[行]高=值
                                ws.column_dimensions['A'].width = 22 #设置第A列的列宽
                                ws.column_dimensions['B'].width = 22
                                ws.column_dimensions['C'].width = 22
                                ws.column_dimensions['D'].width = 22
                                ws.column_dimensions['E'].width = 22
                                ws.column_dimensions['F'].width = 22
                                ws.column_dimensions['G'].width = 40
                                img_C = openpyxl.drawing.image.Image(file) # 插入图片
                                print("img=：", img_C)
                                img_C.width = 115 # 设置图片的宽度
                                img_C.height = 100  # 设置图片的高度
                                ws.add_image(img_C, 'C{}'.format(count))
                                img_C = openpyxl.drawing.image.Image(file)
                                img_D = openpyxl.drawing.image.Image(shotfile_sb)
                                img_D.width = 150  # 设置图片的宽度
                                img_D.height = 100  # 设置图片的高度
                                ws.add_image(img_D, 'D{}'.format(count))
                                img_D = openpyxl.drawing.image.Image(shotfile_sb)
                                print("调试shotfile_dimian的路径：",shotfile_dimian)
                                img_E = openpyxl.drawing.image.Image(shotfile_dimian)
                                img_E.width = 150  # 设置图片的宽度
                                img_E.height = 100 # 设置图片的高度
                                ws.add_image(img_E, 'E{}'.format(count))
                                img_E = openpyxl.drawing.image.Image(shotfile_dimian)
                                img_F = openpyxl.drawing.image.Image(shotfile_dingmian)
                                print("img=：", img_F)
                                img_F.width = 150  # 设置图片的宽度
                                img_F.height = 100  # 设置图片的高度
                                ws.add_image(img_F, 'F{}'.format(count))
                                img_F = openpyxl.drawing.image.Image(shotfile_dingmian)
                                wb.save(r"C:\Users\User\Desktop\3333\99.xlsx")
                                ws = wb.active
                                print('程序结束！')
                                break
                        time.sleep(3)
                        pg.moveTo(x=1142, y=742)
                        pg.click()
                        bo().click_loc(loc.index_page.quxiao_01)
                        for i in range(50):
                            if bo().get_loc(loc.index_page.sheng_cheng_zheng):
                                bo().click_loc(loc.index_page.quxiao_01)
                                pg.moveTo(x=1142, y=742)
                                pg.click()
                                pg.moveTo(x=1342, y=842)
                            else:
                                break
                        bo().click_loc(loc.index_page.x)
                        bo().click_loc(loc.index_page.baocun_tuichu)
                        time.sleep(15)
                        for x in range(10):
                            if bo().get_loc(loc.index_page.xiangmu_guanli):
                                for i in range(10):
                                    time.sleep(2)
                                    if bo().get_loc(loc.index_page.weimingmingfangan):
                                        bo().click_loc(loc.index_page.weimingmingfangan)
                                        time.sleep(2)
                                        pyperclip.copy(f"{name}_{list_i}")
                                        time.sleep(2)
                                        pg.hotkey('ctrl', 'v')
                                        time.sleep(2)
                                        pg.press('enter')
                                    if not bo().get_loc(loc.index_page.weimingmingfangan):
                                        break
                                break
                    time.sleep(5)
                    bo().click_loc(loc.index_page.xiangmu_guanli)
                    bo().click_loc(loc.index_page.xinjian)
                    for i in range(10):
                        if not bo().get_loc(loc.index_page.ziyouhuizhi):
                            bo().click_loc(loc.index_page.xinjian)
                        else:
                            break
                    for i in range(10):
                        if not bo().get_loc(loc.index_page.huxing):
                            bo().click_loc(loc.index_page.ziyouhuizhi)
                        else:
                            break
                    bo().move(loc.index_page.huxing)
                    bo().click_loc(loc.index_page.daorutupian)
                except Exception:
                    logger.info(f"{name[:-4]} 执行失败了！")
                    faillist.append(name[:-5])
        logger.info('执行完毕')
if __name__ == '__main__':
    pass