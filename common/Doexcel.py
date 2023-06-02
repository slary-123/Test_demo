from openpyxl import load_workbook
from openpyxl.styles import Alignment
class Doexcel: # 定义一个表格类
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
    #把测试结果写入excel
    def write_back(self,i,result,Testresult,img): # 将测试结果写回excel
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        sheet.cell(i,1).value=result
        sheet.cell(i,2).value = Testresult
        sheet.cell(i,3).value = img
        Alignment(wrap_text=True)
        wb.save(self.file_name)

if __name__ == '__main__':
    Doexcel(r"C:\Users\user\Desktop\plan.xlsx", "Sheet1").write_back(1,"name","title","text")
