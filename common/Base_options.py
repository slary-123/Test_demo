'''
pyautogui基础操作封装
'''
import traceback
import openpyxl
import pyautogui as pg
from time import sleep
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from common.my_logger import MyLogger
import openpyxl.drawing.image
import xlrd
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
class Base_options:
    def __init__(self): # 进来就等三秒
        sleep(3)
    def click(self, x, y,desc=None): # 单击
        try:
            sleep(1)
            pg.click(x, y)
            sleep(3)
        except Exception:
            print('单击失败')
        else:
            print("执行成功")
        finally:
            print('执行结束')
    def doubleClick(self, x, y,desc=None): # 双击
        sleep(2)
        pg.doubleClick(x, y)
    def mouseClick(self, x, y,desc=None): # 鼠标点击
        sleep(3)
        pg.mouseDown(x=x, y=y)
        sleep(3)
        pg.mouseUp(x=x, y=y)
    @staticmethod
    def get_path(image,desc=None): #获取图像文件的路径
        import os
        image_path = r"C:\Users\user\Desktop\loc_img"
        path = os.path.join(image_path,image) # 将两个路径拼接起来,将图片存放路径和图片名称放在一起
        print("图片路径：",path)
        return path
    def get_loc(self, image,desc=None):
        try:
            png = Base_options.get_path(image)
            a = pg.locateOnScreen(png)
            x, y = pg.center(a)
            print(f'打印出坐标：{x}:{y}')
            return x, y
        except Exception:
            print('坐标获取失败')
    def click_loc(self, image,desc=None):
        try:
            x, y = self.get_loc(image)  # 调用get_loc方法获取图像位置中心坐标
            self.click(x, y)  # 点击坐标
        except Exception:
            print('点击失败')
        else:
            print("执行成功")
        finally:
            print('执行结束')
    def doubleClick_loc(self, image):
        try:
            for i in range(3):
                x, y = self.get_loc(image) #调用get_loc方法获取图像位置中心坐标
                sleep(1)
                self.doubleClick(x, y)#点击坐标
        except Exception:
            print('双击失败')
        else:
            print("执行成功")
        finally:
            print('执行结束')
    def mouseClick_loc(self, image,desc=None): # 鼠标点击图片中心点坐标
        for i in range(3):
            x, y = self.get_loc(image) #调用get_loc方法获取图像位置中心坐标
            sleep(1)
            #self.mouseClick(x, y)#点击坐标
    def move(self, image,desc=None): # 获取图片的坐标并将鼠标移动到对应坐标
        try:
            sleep(3)
            x, y = self.get_loc(image)
            pg.moveTo(x, y) # 将鼠标移动至对应的坐标
        except Exception:
            print('移动失败')
        else:
            print("执行成功")
        finally:
            print('执行结束')
    def upload(self, file_path, desc=None): # 上传图片
        try:
            sleep(2)
            pg.typewrite(file_path)  # 将文件路径（文件名）输入到当前焦点所在的文本框中
            sleep(2)
            pg.hotkey('enter')
            sleep(1)
            pg.hotkey('enter')  # 模拟按下回车键，完成上传操作。
        except Exception:
            print('上传失败')
        else:
            print("执行成功")
        finally:
            print('执行结束')
    def handle_house_file(self,file_path,desc=None):
        import os
        for root, dirs, files in os.walk(file_path): #  r"C:\Users\User\Desktop\big-house" 放户型图的文件
            print("打印路径root:",root)
            print("打印路径dirs:",dirs)
            print("打印路径files:",files)
            return files
    def get_pic_to_list(self,filepath,desc = None):
        import os
        file_path_list = os.listdir(filepath)
        for i in range(len(file_path_list)):
            file_path_list[i] = r"{}\{}".format(filepath,file_path_list[i])
        return file_path_list
    def insert_img(excel_path, img_path, houseID=None, row=None): # 将图片插入表格
            # 写入图片至excel
            # :param excel_path: excel文件地址：C:\Users\User\Desktop\test1.xlsx
            # :param img_path: 图片地址 列表
            # :param houseID:
            # :param row: 写入行数
            # :return:
        wb = Workbook() # 创建一个工作簿
        ws = wb.active # 获取当前激活的sheet表对象
        red_fill = PatternFill(fill_type='solid', start_color='FFC0D9D9',end_color='FFC0D9D9')# 设置表头颜色
        ws.cell(1, 1, "户型ID").fill = red_fill
        ws.cell(1, 2, "原始户型图").fill = red_fill
        ws.cell(1, 3, "户型识别图").fill = red_fill
        ws.cell(1, 4, "彩图缩略图").fill = red_fill
        ws.cell(row, 1, f"{houseID}")  # 指定行列进行追加
        col_ = {'A': 12, 'B': 20, 'C': 20, 'D': 20}
        col_1 = ['B', 'C', 'D']
        ws.row_dimensions[row].height = 80  # ws对象.row_dimensions[行]高=值
        for k, v in col_.items():
            ws.column_dimensions[k].width = v  # 设置第A列的列宽
        for img in range(len(img_path)):
            img_C = openpyxl.drawing.image.Image(img_path[img])
            img_C.width = 120  # 设置图片的宽度
            img_C.height = 100  # 设置图片的高度
            ws.add_image(img_C, f'{col_1[img]}{row}')
        wb.save(excel_path)
        return True
if __name__ == '__main__':
    file = Base_options()
    file_path = r"C:\Users\user\Desktop\house-file"
    file_list = file.handle_house_file(file_path)
    print(f"打印file文件：{file_list}")
